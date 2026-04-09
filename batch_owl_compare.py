#!/usr/bin/env python3
"""
Batch OWL Comparison Tool
Compares multiple extracted OWL files against gold standards.
Generates a comprehensive Excel report with all comparisons.
"""

import sys
import json
from pathlib import Path
from owl_compare import compare_ontologies

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[!] openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def batch_compare(extracted_dir, gold_dir, output_excel):
    """Compare all extracted files against matching gold files."""
    ext_dir = Path(extracted_dir)
    gold_dir = Path(gold_dir)
    
    # Find all gold files
    gold_files = sorted(gold_dir.glob("*_gold.owl"))
    if not gold_files:
        print(f"[!] No *_gold.owl files found in {gold_dir}")
        return
    
    print(f"\nBatch OWL Comparison")
    print(f"  Extracted dir: {extracted_dir}")
    print(f"  Gold dir:      {gold_dir}")
    print(f"  Gold files:    {len(gold_files)}\n")
    
    results = []
    
    for gold_file in gold_files:
        stem = gold_file.stem.replace("_gold", "")
        
        # Find matching extracted file
        extracted_file = ext_dir / stem / f"{stem}_ontology.owl"
        
        if not extracted_file.exists():
            print(f"[!] Skipping {stem}: extracted file not found at {extracted_file}")
            continue
        
        print(f"Comparing {stem}...")
        comparison = compare_ontologies(str(extracted_file), str(gold_file))
        
        if not comparison:
            print(f"  [ERROR] Comparison failed for {stem}")
            continue
        
        comparison["file_stem"] = stem
        comparison["extracted_file"] = str(extracted_file)
        comparison["gold_file"] = str(gold_file)
        results.append(comparison)
        
        m = comparison["metrics"]
        print(f"  F1: Classes={m['classes_f1']:.1%} OProps={m['oprops_f1']:.1%} "
              f"DProps={m['dprops_f1']:.1%} Overall={m['overall_f1']:.1%}")
    
    if not results:
        print("[!] No comparisons were successful")
        return
    
    print(f"\n✓ {len(results)} comparisons complete. Saving Excel report...")
    save_batch_excel(results, output_excel)
    print(f"✓ Report saved to: {output_excel}\n")


def save_batch_excel(results, output_file):
    """Save all comparisons to a multi-sheet Excel workbook."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Define colors
    TITLE_FILL = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
    HDR_FILL   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    GOOD_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    BAD_FILL   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    WARN_FILL  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def hdr_cell(ws, row, col, value):
        c = ws.cell(row=row, column=col)
        c.value = value
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
        return c
    
    def pct_cell(ws, row, col, value):
        c = ws.cell(row=row, column=col)
        c.value = value
        c.number_format = '0.0%'
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border
        if value >= 0.99:
            c.fill = GOOD_FILL
        elif value < 0.5:
            c.fill = BAD_FILL
        elif value < 0.9:
            c.fill = WARN_FILL
        return c
    
    # ── Sheet 1: Summary Table ──────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary", 0)
    
    col_widths = [15, 12, 12, 12, 12, 12, 12]
    for col_idx, width in enumerate(col_widths, 1):
        ws_summary.column_dimensions[chr(64 + col_idx)].width = width
    
    # Title
    ws_summary.merge_cells("A1:G1")
    title = ws_summary.cell(row=1, column=1)
    title.value = "OWL Comparison Summary"
    title.font = TITLE_FONT
    title.fill = TITLE_FILL
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 22
    
    # Headers
    row = 3
    headers = ["File Stem", "Classes F1", "ObjProps F1", "DataProps F1", 
               "Restrictions F1", "Overall F1", "Status"]
    for col, header in enumerate(headers, 1):
        hdr_cell(ws_summary, row, col, header)
    
    # Data rows
    row = 4
    for result in sorted(results, key=lambda r: r["file_stem"]):
        stem = result["file_stem"]
        m = result["metrics"]
        
        ws_summary.cell(row=row, column=1).value = stem
        ws_summary.cell(row=row, column=1).border = thin_border
        
        pct_cell(ws_summary, row, 2, m["classes_f1"])
        pct_cell(ws_summary, row, 3, m["oprops_f1"])
        pct_cell(ws_summary, row, 4, m["dprops_f1"])
        pct_cell(ws_summary, row, 5, m["restrictions_f1"])
        
        overall_f1 = m["overall_f1"]
        c = pct_cell(ws_summary, row, 6, overall_f1)
        
        # Status based on overall F1
        status_cell = ws_summary.cell(row=row, column=7)
        if overall_f1 >= 0.95:
            status_cell.value = "EXCELLENT"
            status_cell.fill = GOOD_FILL
        elif overall_f1 >= 0.8:
            status_cell.value = "GOOD"
            status_cell.fill = GOOD_FILL
        elif overall_f1 >= 0.6:
            status_cell.value = "OK"
            status_cell.fill = WARN_FILL
        else:
            status_cell.value = "NEEDS REVIEW"
            status_cell.fill = BAD_FILL
        status_cell.alignment = Alignment(horizontal="center")
        status_cell.border = thin_border
        
        row += 1
    
    # Averages row
    row += 1
    avg_c_f1 = sum(r["metrics"]["classes_f1"] for r in results) / len(results)
    avg_o_f1 = sum(r["metrics"]["oprops_f1"] for r in results) / len(results)
    avg_d_f1 = sum(r["metrics"]["dprops_f1"] for r in results) / len(results)
    avg_r_f1 = sum(r["metrics"]["restrictions_f1"] for r in results) / len(results)
    avg_overall_f1 = sum(r["metrics"]["overall_f1"] for r in results) / len(results)
    
    ws_summary.cell(row=row, column=1).value = "AVERAGE"
    ws_summary.cell(row=row, column=1).font = Font(bold=True)
    ws_summary.cell(row=row, column=1).border = thin_border
    pct_cell(ws_summary, row, 2, avg_c_f1)
    pct_cell(ws_summary, row, 3, avg_o_f1)
    pct_cell(ws_summary, row, 4, avg_d_f1)
    pct_cell(ws_summary, row, 5, avg_r_f1)
    pct_cell(ws_summary, row, 6, avg_overall_f1)
    
    # ── Detail Sheets (one per file) ────────────────────────────────────────
    for sheet_idx, result in enumerate(sorted(results, key=lambda r: r["file_stem"]), 1):
        stem = result["file_stem"]
        sheet_name = stem[:31]  # Excel sheet name limit
        ws = wb.create_sheet(sheet_name, sheet_idx)
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 6
        ws.column_dimensions['E'].width = 6
        ws.column_dimensions['F'].width = 6
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        
        row = 1
        # Title
        ws.merge_cells(f'A{row}:I{row}')
        title_cell = ws.cell(row=row, column=1)
        title_cell.value = f"Comparison: {stem}"
        title_cell.font = TITLE_FONT
        title_cell.fill = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20
        
        # Element metrics table
        row = 3
        headers = ["Element", "Gold", "Ext", "TP", "FP", "FN", "Prec", "Rec", "F1"]
        for col, header in enumerate(headers, 1):
            hdr_cell(ws, row, col, header)
        
        # Element rows
        row = 4
        elements = [
            ("classes", "Classes"),
            ("object_properties", "Object Properties"),
            ("data_properties", "Data Properties"),
            ("restrictions", "Restrictions"),
        ]
        
        for elem_key, elem_label in elements:
            ws.cell(row=row, column=1).value = elem_label
            ws.cell(row=row, column=1).border = thin_border
            
            if elem_key == "restrictions":
                r = result["restrictions"]
                g = r["gold_count"]
                e = r["extracted_count"]
                tp = r["correct_count"]
                fp = r["extra_count"]
                fn = r["missing_count"]
                m = result["metrics"]
                prec = m["restrictions_precision"]
                rec = m["restrictions_recall"]
                f1 = m["restrictions_f1"]
            else:
                c = result[elem_key]
                g = len(c["gold"])
                e = len(c["extracted"])
                tp = len(c["correct"])
                fp = len(c["extra"])
                fn = len(c["missing"])
                m = result["metrics"]
                prefix = {"classes": "classes", "object_properties": "oprops",
                          "data_properties": "dprops"}[elem_key]
                prec = m[f"{prefix}_precision"]
                rec = m[f"{prefix}_recall"]
                f1 = m[f"{prefix}_f1"]
            
            for col, val in enumerate([g, e, tp, fp, fn], 2):
                c = ws.cell(row=row, column=col)
                c.value = val
                c.alignment = Alignment(horizontal="center")
                c.border = thin_border
            
            pct_cell(ws, row, 7, prec)
            pct_cell(ws, row, 8, rec)
            pct_cell(ws, row, 9, f1)
            
            row += 1
        
        # Mismatch notes section
        row += 2
        ws.merge_cells(f'A{row}:I{row}')
        notes_title = ws.cell(row=row, column=1)
        notes_title.value = "MISMATCH NOTES"
        notes_title.font = Font(bold=True, size=11)
        notes_title.fill = WARN_FILL
        
        row += 1
        notes = result["mismatch_notes"]
        for note in notes:
            if note["severity"] != "NONE":
                ws.merge_cells(f'A{row}:I{row}')
                c = ws.cell(row=row, column=1)
                c.value = (f"[{note['severity']}] {note['category']}: {note['element']}")
                c.font = Font(bold=True)
                if note["severity"] == "FALSE NEGATIVE":
                    c.fill = BAD_FILL
                elif note["severity"] == "FALSE POSITIVE":
                    c.fill = WARN_FILL
                c.alignment = Alignment(wrap_text=True)
                row += 1
                
                ws.merge_cells(f'A{row}:I{row}')
                c = ws.cell(row=row, column=1)
                c.value = note["note"]
                c.alignment = Alignment(wrap_text=True)
                ws.row_dimensions[row].height = 30
                row += 2
    
    wb.save(output_file)


if __name__ == "__main__":
    extracted = r"f:\Ontology extractor codes\evaluation_toolkit\evaluation_toolkit\results_real_images"
    gold = r"f:\Ontology extractor codes\evaluation_toolkit\evaluation_toolkit\results\owl_gold_real_images"
    output = r"f:\Ontology extractor codes\evaluation_toolkit\evaluation_toolkit\OWL_COMPARISON_REPORT.xlsx"
    
    batch_compare(extracted, gold, output)
