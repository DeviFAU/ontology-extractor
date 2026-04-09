"""
evaluate.py
===========
Compares ontology extractor output against gold-standard JSON.
Produces per-element precision/recall/F1 and Edge Direction Accuracy (EDA).

Usage:
  python evaluate.py --extraction result.json --gold gold.json
  python evaluate.py --extraction result.json --gold gold.json --output report.json
  python evaluate.py --extraction result.json --gold gold.json --output-excel report.xlsx
  python evaluate.py --batch-extraction results/ --batch-gold gold_standards/ --output report.json

Changes vs original:
  - restriction_axioms now included in overall F1 (was excluded before).
  - --output-excel flag: saves full per-element results to Excel.

Requirements: openpyxl (only for Excel output)
"""

import json
import argparse
import csv
from pathlib import Path
from collections import defaultdict


# ─── Normalisation ───────────────────────────────────────────────────────────

def _norm(name: str) -> str:
    """Normalise a name for comparison: lowercase, strip prefix, strip spaces."""
    n = name.lower().strip().replace("_", " ")
    # Strip known prefixes
    for prefix in ["e:", "prov:", "foaf:", "schema:", "rdfs:", "owl:", "rdf:",
                    "http://example.org/ontology#",
                    "http://example.org/", "http://e.org/ont#"]:
        if n.startswith(prefix):
            n = n[len(prefix):]
    return n.strip()


def _edge_key(edge: dict) -> tuple:
    """Create a normalised tuple for edge comparison."""
    return (
        _norm(edge.get("from", "")),
        _norm(edge.get("to", "")),
        _norm(edge.get("label", "")),
        _norm(edge.get("edge_type", "")),
    )


def _edge_key_no_direction(edge: dict) -> tuple:
    """Edge key ignoring direction (for EDA computation)."""
    f, t = _norm(edge.get("from", "")), _norm(edge.get("to", ""))
    return (min(f, t), max(f, t), _norm(edge.get("label", "")))


def _edge_direction_key(edge: dict) -> tuple:
    """Edge key with direction for EDA."""
    return (
        _norm(edge.get("from", "")),
        _norm(edge.get("to", "")),
        _norm(edge.get("label", "")),
    )


# ─── Element-Level Evaluation ────────────────────────────────────────────────

def evaluate_set(gold_items: list, extracted_items: list,
                 key_fn=None) -> dict:
    """Compare two lists using a key function. Returns TP, FP, FN, P, R, F1."""
    if key_fn is None:
        key_fn = lambda x: _norm(x) if isinstance(x, str) else _norm(str(x))

    gold_keys = set(key_fn(item) for item in gold_items)
    ext_keys = set(key_fn(item) for item in extracted_items)

    tp = gold_keys & ext_keys
    fp = ext_keys - gold_keys
    fn = gold_keys - ext_keys

    precision = len(tp) / len(ext_keys) if ext_keys else 1.0
    recall = len(tp) / len(gold_keys) if gold_keys else 1.0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0.0

    return {
        "gold_count": len(gold_keys),
        "extracted_count": len(ext_keys),
        "tp": len(tp),
        "fp": len(fp),
        "fn": len(fn),
        "precision": round(precision, 4),
        "recall": round(recall, 4),
        "f1": round(f1, 4),
        "false_positives": sorted(list(fp))[:10],
        "false_negatives": sorted(list(fn))[:10],
    }


# ─── Edge Direction Accuracy ─────────────────────────────────────────────────

def compute_eda(gold_edges: list, extracted_edges: list) -> dict:
    """
    Edge Direction Accuracy: Of edges where both systems found the same
    two nodes + label, what fraction have the correct direction?
    """
    gold_directed = {_edge_direction_key(e) for e in gold_edges}
    ext_directed = {_edge_direction_key(e) for e in extracted_edges}

    gold_undirected = {_edge_key_no_direction(e) for e in gold_edges}
    ext_undirected = {_edge_key_no_direction(e) for e in extracted_edges}

    # Edges found by both (same nodes + label, ignoring direction)
    common_undirected = gold_undirected & ext_undirected

    if not common_undirected:
        return {
            "common_edges": 0,
            "correctly_directed": 0,
            "eda": 0.0,
            "reversed_edges": [],
        }

    # Of these common edges, how many have correct direction?
    correctly_directed = 0
    reversed_edges = []
    for ue in common_undirected:
        # Find the gold and extracted directed versions
        g_matches = [e for e in gold_edges
                     if _edge_key_no_direction(e) == ue]
        e_matches = [e for e in extracted_edges
                     if _edge_key_no_direction(e) == ue]
        if g_matches and e_matches:
            g_dir = _edge_direction_key(g_matches[0])
            e_dir = _edge_direction_key(e_matches[0])
            if g_dir == e_dir:
                correctly_directed += 1
            else:
                reversed_edges.append({
                    "label": g_matches[0].get("label", ""),
                    "gold": f"{g_matches[0]['from']} -> {g_matches[0]['to']}",
                    "extracted": f"{e_matches[0]['from']} -> {e_matches[0]['to']}",
                })

    eda = correctly_directed / len(common_undirected) if common_undirected else 0.0

    return {
        "common_edges": len(common_undirected),
        "correctly_directed": correctly_directed,
        "eda": round(eda, 4),
        "reversed_edges": reversed_edges[:10],
    }


# ─── Full Evaluation ─────────────────────────────────────────────────────────

def evaluate(gold: dict, extraction: dict) -> dict:
    """Full evaluation of extraction against gold standard."""

    # Get the final merged extraction (may be nested under 'final')
    ext = extraction.get("final", extraction)

    # ── Classes ──────────────────────────────────────────────────────────
    gold_classes = [c["name"] for c in gold.get("classes", [])]
    ext_classes = [c["name"] for c in ext.get("classes", [])]
    class_eval = evaluate_set(gold_classes, ext_classes)

    # ── Object Properties ────────────────────────────────────────────────
    gold_oprops = [op["name"] for op in gold.get("object_properties", [])]
    ext_oprops = [op["name"] for op in ext.get("object_properties", [])]
    oprop_eval = evaluate_set(gold_oprops, ext_oprops)

    # ── Data Properties ──────────────────────────────────────────────────
    gold_dprops = [dp["name"] if isinstance(dp, dict) and "name" in dp
                   else dp.get("property", "")
                   for dp in gold.get("data_properties", [])]
    ext_dprops = [dp.get("property", dp.get("name", ""))
                  for dp in ext.get("data_properties", [])]
    dprop_eval = evaluate_set(gold_dprops, ext_dprops)

    # ── Instances ────────────────────────────────────────────────────────
    gold_instances = [i["name"] for i in gold.get("instances", [])]
    ext_instances = [i["name"] for i in ext.get("instances", [])]
    instance_eval = evaluate_set(gold_instances, ext_instances)

    # ── Edges (full: from+to+label+type) ─────────────────────────────────
    gold_edges = gold.get("edges", [])
    ext_edges = ext.get("edges", [])
    edge_eval = evaluate_set(gold_edges, ext_edges, key_fn=_edge_key)

    # ── Edges by type ────────────────────────────────────────────────────
    edge_types = set()
    for e in gold_edges + ext_edges:
        edge_types.add(e.get("edge_type", "other"))

    per_type = {}
    for etype in sorted(edge_types):
        g_typed = [e for e in gold_edges if e.get("edge_type") == etype]
        e_typed = [e for e in ext_edges if e.get("edge_type") == etype]
        per_type[etype] = evaluate_set(g_typed, e_typed, key_fn=_edge_key)

    # ── Edge Direction Accuracy ──────────────────────────────────────────
    eda = compute_eda(gold_edges, ext_edges)

    # ── Restriction Axioms ───────────────────────────────────────────────
    def _rx_key(rx):
        return (_norm(rx.get("subject", "")),
                _norm(rx.get("property", "")),
                _norm(rx.get("restriction_type", "")),
                _norm(rx.get("filler", "")))
    gold_rx = gold.get("restriction_axioms", [])
    ext_rx = ext.get("restriction_axioms", [])
    rx_eval = evaluate_set(gold_rx, ext_rx, key_fn=_rx_key)

    # ── Data Assertions ──────────────────────────────────────────────────
    def _da_key(da):
        return (_norm(da.get("individual", "")),
                _norm(da.get("property", "")),
                da.get("value", "").lower().strip())
    gold_da = gold.get("data_assertions", [])
    ext_da = ext.get("data_assertions", [])
    da_eval = evaluate_set(gold_da, ext_da, key_fn=_da_key)

    # ── Overall F1 (weighted by element count) ───────────────────────────
    # NOTE: restriction_axioms are included so the headline score reflects
    # axioms that are in the gold OWL but invisible in the diagram.
    components = [
        (class_eval,    "classes"),
        (oprop_eval,    "object_properties"),
        (dprop_eval,    "data_properties"),
        (instance_eval, "instances"),
        (edge_eval,     "edges"),
        (rx_eval,       "restriction_axioms"),   # ← added
    ]
    total_tp = sum(c[0]["tp"] for c in components)
    total_fp = sum(c[0]["fp"] for c in components)
    total_fn = sum(c[0]["fn"] for c in components)
    overall_p = total_tp / (total_tp + total_fp) if (total_tp + total_fp) > 0 else 0
    overall_r = total_tp / (total_tp + total_fn) if (total_tp + total_fn) > 0 else 0
    overall_f1 = 2 * overall_p * overall_r / (overall_p + overall_r) \
                 if (overall_p + overall_r) > 0 else 0

    return {
        "gold_id": gold.get("id", "unknown"),
        "gold_name": gold.get("name", "unknown"),
        "complexity": gold.get("complexity", "unknown"),
        "diagram_type": gold.get("diagram_type", "unknown"),
        "classes": class_eval,
        "object_properties": oprop_eval,
        "data_properties": dprop_eval,
        "instances": instance_eval,
        "edges": edge_eval,
        "edges_by_type": per_type,
        "edge_direction_accuracy": eda,
        "restriction_axioms": rx_eval,
        "data_assertions": da_eval,
        "overall": {
            "tp": total_tp, "fp": total_fp, "fn": total_fn,
            "precision": round(overall_p, 4),
            "recall": round(overall_r, 4),
            "f1": round(overall_f1, 4),
        },
    }


# ─── Report Printer ──────────────────────────────────────────────────────────

def print_report(result: dict):
    """Print a formatted evaluation report."""
    print(f"\n{'='*70}")
    print(f"  Evaluation: {result['gold_name']} ({result['gold_id']})")
    print(f"  Complexity: {result['complexity']}  |  Type: {result['diagram_type']}")
    print(f"{'='*70}")

    print(f"\n{'Element':<25} {'Gold':>5} {'Ext':>5} {'TP':>4} {'FP':>4} {'FN':>4} "
          f"{'Prec':>7} {'Rec':>7} {'F1':>7}")
    print(f"{'-'*70}")

    for key in ["classes", "object_properties", "data_properties",
                "instances", "edges", "restriction_axioms", "data_assertions"]:
        r = result[key]
        print(f"{key:<25} {r['gold_count']:>5} {r['extracted_count']:>5} "
              f"{r['tp']:>4} {r['fp']:>4} {r['fn']:>4} "
              f"{r['precision']:>7.1%} {r['recall']:>7.1%} {r['f1']:>7.1%}")

    # Edges by type
    if result.get("edges_by_type"):
        print(f"\n  Edge breakdown by type:")
        for etype, r in result["edges_by_type"].items():
            if r["gold_count"] > 0 or r["extracted_count"] > 0:
                print(f"    {etype:<22} gold={r['gold_count']}  ext={r['extracted_count']}  "
                      f"TP={r['tp']}  F1={r['f1']:.1%}")

    # EDA
    eda = result["edge_direction_accuracy"]
    print(f"\n  Edge Direction Accuracy: {eda['eda']:.1%} "
          f"({eda['correctly_directed']}/{eda['common_edges']} common edges)")
    if eda["reversed_edges"]:
        print(f"  Reversed edges:")
        for rev in eda["reversed_edges"]:
            print(f"    {rev['label']}: gold={rev['gold']}  extracted={rev['extracted']}")

    # Overall
    o = result["overall"]
    print(f"\n  OVERALL: P={o['precision']:.1%}  R={o['recall']:.1%}  F1={o['f1']:.1%}")
    print(f"{'='*70}\n")


# ─── Batch Evaluation ────────────────────────────────────────────────────────

def batch_evaluate(extraction_dir: str, gold_dir: str, output_path: str = None, output_excel_path: str = None):
    """Evaluate all extraction JSONs against matching gold JSONs."""
    ext_path = Path(extraction_dir)
    gold_path = Path(gold_dir)

    # Match by stem: synth_01_extraction.json <-> synth_01_gold.json
    gold_files = {gf.stem.replace("_gold", ""): gf
                  for gf in gold_path.glob("*_gold.json")}

    ext_files = {}
    for ef in ext_path.rglob("*_extraction.json"):
        stem = ef.stem.replace("_extraction", "")
        ext_files[stem] = ef

    matched = set(gold_files.keys()) & set(ext_files.keys())
    if not matched:
        print(f"[!] No matching files found between {extraction_dir} and {gold_dir}")
        print(f"    Gold stems: {sorted(gold_files.keys())}")
        print(f"    Extraction stems: {sorted(ext_files.keys())}")
        return

    print(f"\nEvaluating {len(matched)} matched pairs...")
    results = []

    for stem in sorted(matched):
        with open(gold_files[stem]) as f:
            gold = json.load(f)
        with open(ext_files[stem]) as f:
            extraction = json.load(f)
        result = evaluate(gold, extraction)
        results.append(result)
        print_report(result)

    # Summary table
    print(f"\n{'='*90}")
    print(f"  SUMMARY")
    print(f"{'='*90}")
    print(f"{'ID':<15} {'Name':<25} {'Cmplx':<8} {'Cls F1':>7} {'Edge F1':>8} "
          f"{'EDA':>6} {'Overall':>8}")
    print(f"{'-'*90}")

    for r in results:
        print(f"{r['gold_id']:<15} {r['gold_name']:<25} {r['complexity']:<8} "
              f"{r['classes']['f1']:>7.1%} {r['edges']['f1']:>8.1%} "
              f"{r['edge_direction_accuracy']['eda']:>6.1%} {r['overall']['f1']:>8.1%}")

    # Averages
    avg_cls = sum(r["classes"]["f1"] for r in results) / len(results)
    avg_edge = sum(r["edges"]["f1"] for r in results) / len(results)
    avg_eda = sum(r["edge_direction_accuracy"]["eda"] for r in results) / len(results)
    avg_overall = sum(r["overall"]["f1"] for r in results) / len(results)

    print(f"{'-'*90}")
    print(f"{'AVERAGE':<15} {'':<25} {'':<8} {avg_cls:>7.1%} {avg_edge:>8.1%} "
          f"{avg_eda:>6.1%} {avg_overall:>8.1%}")
    print(f"{'='*90}\n")

    # Save JSON report
    if output_path:
        report = {
            "results": results,
            "summary": {
                "count": len(results),
                "avg_class_f1": round(avg_cls, 4),
                "avg_edge_f1": round(avg_edge, 4),
                "avg_eda": round(avg_eda, 4),
                "avg_overall_f1": round(avg_overall, 4),
            }
        }
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2, ensure_ascii=False)
        print(f"Report saved to: {output_path}")

    # Also save CSV
    if output_path:
        csv_path = Path(output_path).with_suffix(".csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "Name", "Complexity", "Type",
                            "Class_P", "Class_R", "Class_F1",
                            "Edge_P", "Edge_R", "Edge_F1",
                            "EDA", "Overall_F1"])
            for r in results:
                writer.writerow([
                    r["gold_id"], r["gold_name"], r["complexity"], r["diagram_type"],
                    r["classes"]["precision"], r["classes"]["recall"], r["classes"]["f1"],
                    r["edges"]["precision"], r["edges"]["recall"], r["edges"]["f1"],
                    r["edge_direction_accuracy"]["eda"], r["overall"]["f1"],
                ])
        print(f"CSV saved to: {csv_path}")
    
    # Save detailed Excel
    if output_excel_path:
        save_batch_excel(results, output_excel_path)

    # Also save CSV
    if output_path:
        csv_path = Path(output_path).with_suffix(".csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "Name", "Complexity", "Type",
                            "Class_P", "Class_R", "Class_F1",
                            "Edge_P", "Edge_R", "Edge_F1",
                            "EDA", "Overall_F1"])
            for r in results:
                writer.writerow([
                    r["gold_id"], r["gold_name"], r["complexity"], r["diagram_type"],
                    r["classes"]["precision"], r["classes"]["recall"], r["classes"]["f1"],
                    r["edges"]["precision"], r["edges"]["recall"], r["edges"]["f1"],
                    r["edge_direction_accuracy"]["eda"], r["overall"]["f1"],
                ])
        print(f"CSV saved to: {csv_path}")


def save_batch_excel(results: list, path: str):
    """Save detailed batch evaluation results to Excel with exact terminal format."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[!] openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    TITLE_FILL = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
    HDR_FILL  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    BAD_FILL  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ── Create one sheet per example with exact terminal format ──────────
    for result in results:
        sheet_name = result['gold_id'][:31]  # Excel sheet names max 31 chars
        ws = wb.create_sheet(sheet_name)
        
        # Set column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 6
        ws.column_dimensions['C'].width = 6
        ws.column_dimensions['D'].width = 6
        ws.column_dimensions['E'].width = 6
        ws.column_dimensions['F'].width = 6
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10

        row = 1
        # Title
        ws.merge_cells(f'A{row}:I{row}')
        title = ws.cell(row=row, column=1)
        title.value = f"Evaluation: {result['gold_name']} ({result['gold_id']})"
        title.font = TITLE_FONT
        title.fill = TITLE_FILL
        title.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 2
        ws.merge_cells(f'A{row}:I{row}')
        subtitle = ws.cell(row=row, column=1)
        subtitle.value = f"Complexity: {result['complexity']}  |  Type: {result['diagram_type']}"
        subtitle.font = Font(size=10)
        subtitle.alignment = Alignment(horizontal="center")
        
        # Element table header
        row = 4
        headers = ["Element", "Gold", "Ext", "TP", "FP", "FN", "Prec", "Rec", "F1"]
        for col, header in enumerate(headers, 1):
            c = ws.cell(row=row, column=col)
            c.value = header
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
        
        # Element rows
        row = 5
        elements = ["classes", "object_properties", "data_properties", "instances", 
                   "edges", "restriction_axioms", "data_assertions"]
        element_names = ["Classes", "Object Properties", "Data Properties", "Instances",
                        "Edges", "Restriction Axioms", "Data Assertions"]
        
        for elem_key, elem_name in zip(elements, element_names):
            e = result[elem_key]
            c = ws.cell(row=row, column=1)
            c.value = elem_name
            c.border = thin_border
            
            ws.cell(row=row, column=2).value = e['gold_count']
            ws.cell(row=row, column=3).value = e['extracted_count']
            ws.cell(row=row, column=4).value = e['tp']
            ws.cell(row=row, column=5).value = e['fp']
            ws.cell(row=row, column=6).value = e['fn']
            ws.cell(row=row, column=7).value = e['precision']
            ws.cell(row=row, column=8).value = e['recall']
            ws.cell(row=row, column=9).value = e['f1']
            
            # Format numbers and colors
            for col in range(2, 10):
                c = ws.cell(row=row, column=col)
                c.alignment = Alignment(horizontal="center")
                c.border = thin_border
                if col >= 7:  # Percentage columns
                    c.number_format = '0.0%'
                    if col == 9:  # F1 column - color code
                        if e['f1'] >= 0.99:
                            c.fill = GOOD_FILL
                        elif e['f1'] < 0.5:
                            c.fill = BAD_FILL
                        elif e['f1'] < 0.9:
                            c.fill = WARN_FILL
            
            row += 1
        
        # Edge breakdown by type
        row += 2
        ws.merge_cells(f'A{row}:I{row}')
        edge_title = ws.cell(row=row, column=1)
        edge_title.value = "Edge breakdown by type:"
        edge_title.font = Font(bold=True, size=10)
        
        row += 1
        edge_headers = ["Edge Type", "Gold", "Ext", "TP", "F1"]
        for col, header in enumerate(edge_headers, 1):
            c = ws.cell(row=row, column=col)
            c.value = header
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border
        
        row += 1
        if result.get("edges_by_type"):
            for etype, edata in result["edges_by_type"].items():
                if edata["gold_count"] > 0 or edata["extracted_count"] > 0:
                    ws.cell(row=row, column=1).value = etype
                    ws.cell(row=row, column=2).value = edata['gold_count']
                    ws.cell(row=row, column=3).value = edata['extracted_count']
                    ws.cell(row=row, column=4).value = edata['tp']
                    ws.cell(row=row, column=5).value = edata['f1']
                    
                    for col in range(1, 6):
                        c = ws.cell(row=row, column=col)
                        c.alignment = Alignment(horizontal="center")
                        c.border = thin_border
                        if col == 5:
                            c.number_format = '0.0%'
                            if edata['f1'] >= 0.99:
                                c.fill = GOOD_FILL
                            elif edata['f1'] < 0.5:
                                c.fill = BAD_FILL
                    
                    row += 1
        
        # EDA
        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        eda_cell = ws.cell(row=row, column=1)
        eda = result["edge_direction_accuracy"]
        eda_cell.value = f"Edge Direction Accuracy: {eda['eda']:.1%} ({eda['correctly_directed']}/{eda['common_edges']} common edges)"
        eda_cell.font = Font(bold=True, size=10)
        
        # Overall
        row += 2
        ws.merge_cells(f'A{row}:I{row}')
        overall_title = ws.cell(row=row, column=1)
        o = result["overall"]
        overall_title.value = f"OVERALL: P={o['precision']:.1%}  R={o['recall']:.1%}  F1={o['f1']:.1%}"
        overall_title.font = Font(bold=True, size=11, color="FFFFFF")
        overall_title.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        overall_title.alignment = Alignment(horizontal="center")

    wb.save(path)
    print(f"  Excel (Detailed - Terminal Format): {path}")


# ─── Excel Export ────────────────────────────────────────────────────────────

def save_excel(result: dict, path: str):
    """Save a single evaluation result to a formatted Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[!] openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR_FILL  = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HDR_FONT  = Font(bold=True, color="FFFFFF")
    GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    BAD_FILL  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    def _hdr(ws, row, cols):
        for col, val in enumerate(cols, 1):
            c = ws.cell(row=row, column=col)
            c.value, c.fill, c.font = val, HDR_FILL, HDR_FONT
            c.alignment = Alignment(horizontal="center")

    def _pct(ws, row, col, val):
        c = ws.cell(row=row, column=col)
        c.value, c.number_format = val, '0.00%'
        c.alignment = Alignment(horizontal="center")
        if val >= 0.999:   c.fill = GOOD_FILL
        elif val == 0.0:   c.fill = BAD_FILL
        elif val < 0.75:   c.fill = WARN_FILL

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws = wb.create_sheet("Summary", 0)
    for col, w in zip("ABCDEFGHI", [28, 8, 11, 6, 6, 6, 11, 9, 9]):
        ws.column_dimensions[col].width = w
    _hdr(ws, 1, ["Element", "Gold", "Extracted", "TP", "FP", "FN",
                 "Precision", "Recall", "F1"])

    keys = ["classes", "object_properties", "data_properties",
            "instances", "edges", "restriction_axioms", "data_assertions"]
    labels = ["Classes", "Object Properties", "Data Properties",
              "Instances", "Edges", "Restriction Axioms", "Data Assertions"]
    for row_i, (k, lbl) in enumerate(zip(keys, labels), 2):
        r = result[k]
        ws.cell(row=row_i, column=1).value = lbl
        for col, val in enumerate([r["gold_count"], r["extracted_count"],
                                    r["tp"], r["fp"], r["fn"]], 2):
            ws.cell(row=row_i, column=col).value = val
            ws.cell(row=row_i, column=col).alignment = Alignment(horizontal="center")
        _pct(ws, row_i, 7, r["precision"])
        _pct(ws, row_i, 8, r["recall"])
        _pct(ws, row_i, 9, r["f1"])

    row_i += 1
    ws.cell(row=row_i, column=1).value = "OVERALL (incl. restrictions)"
    ws.cell(row=row_i, column=1).font  = Font(bold=True)
    o = result["overall"]
    for col, val in enumerate([o["tp"], o["fp"], o["fn"]], 4):
        ws.cell(row=row_i, column=col).value = val
        ws.cell(row=row_i, column=col).alignment = Alignment(horizontal="center")
    _pct(ws, row_i, 7, o["precision"])
    _pct(ws, row_i, 8, o["recall"])
    _pct(ws, row_i, 9, o["f1"])

    # ── Sheet 2: EDA ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Edge Direction Accuracy", 1)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    _hdr(ws, 1, ["Label", "Gold direction", "Extracted direction", "Status"])
    eda = result["edge_direction_accuracy"]
    ws.cell(row=2, column=1).value = "EDA Score"
    _pct(ws, 2, 2, eda["eda"])
    ws.cell(row=3, column=1).value = f"Correct: {eda['correctly_directed']} / {eda['common_edges']}"
    row = 5
    for rev in eda.get("reversed_edges", []):
        ws.cell(row=row, column=1).value = rev["label"]
        ws.cell(row=row, column=2).value = rev["gold"]
        ws.cell(row=row, column=3).value = rev["extracted"]
        ws.cell(row=row, column=4).value = "REVERSED"
        ws.cell(row=row, column=4).fill  = BAD_FILL
        row += 1

    # ── Sheet 3: Mismatch Notes ───────────────────────────────────────────
    ws = wb.create_sheet("Mismatch Notes", 2)
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    _hdr(ws, 1, ["Element Type", "Issue", "Item(s)"])
    row = 2
    FN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    FP_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for k, lbl in zip(keys, labels):
        r = result[k]
        fns = r.get("false_negatives", [])
        fps = r.get("false_positives", [])
        for item in fns:
            ws.cell(row=row, column=1).value = lbl
            c = ws.cell(row=row, column=2)
            c.value = "FALSE NEGATIVE"
            c.fill  = FN_FILL
            ws.cell(row=row, column=3).value = str(item)
            row += 1
        for item in fps:
            ws.cell(row=row, column=1).value = lbl
            c = ws.cell(row=row, column=2)
            c.value = "FALSE POSITIVE"
            c.fill  = FP_FILL
            ws.cell(row=row, column=3).value = str(item)
            row += 1
    if row == 2:
        ws.cell(row=2, column=1).value = "No mismatches found"
        ws.cell(row=2, column=1).fill  = GOOD_FILL

    wb.save(path)
    print(f"  Excel : {path}")


# ─── CLI ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Evaluate ontology extraction against gold standard")
    parser.add_argument("--extraction", "-e", help="Extraction JSON file")
    parser.add_argument("--gold", "-g", help="Gold standard JSON file")
    parser.add_argument("--batch-extraction", help="Directory with extraction JSONs")
    parser.add_argument("--batch-gold", help="Directory with gold JSONs")
    parser.add_argument("--output", "-o", help="Output report path (JSON + CSV)")
    parser.add_argument("--output-excel", "-x", help="Output Excel report path (.xlsx)")
    args = parser.parse_args()

    if args.extraction and args.gold:
        with open(args.gold) as f:
            gold = json.load(f)
        with open(args.extraction) as f:
            ext = json.load(f)
        result = evaluate(gold, ext)
        print_report(result)
        if args.output:
            with open(args.output, "w") as f:
                json.dump(result, f, indent=2)
        if args.output_excel:
            save_excel(result, args.output_excel)
    elif args.batch_extraction and args.batch_gold:
        batch_evaluate(args.batch_extraction, args.batch_gold, args.output, args.output_excel)
    else:
        parser.print_help()
